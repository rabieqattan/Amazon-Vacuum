import requests
import time
import csv
import os
import sys
import threading
import subprocess
import platform
import config
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Hard wall-clock cap on a single Selenium page extraction. Selenium's own
# page_load_timeout doesn't help if chromedriver/Chrome itself wedges (e.g. a
# crashed renderer chromedriver never notices) -- the HTTP call to chromedriver
# can then hang forever with no timeout of its own. Running the extraction in
# a daemon thread lets us give up after this many seconds and force-kill the
# stuck browser instead of blocking the whole run.
SELENIUM_TIMEOUT = 45

class AmazonVacuumCollector:
    def __init__(self, api_key, csv_file_path):
        self.api_key = api_key
        self.csv_file_path = csv_file_path
        self.asins = []
        self.all_products = []
        self.base_url = "https://api.rainforestapi.com/request"

        # Initialize Selenium for seller info extraction
        self.driver = None
        self._init_selenium()
        self._load_asins_from_csv()
        # Fixed at load time so the Summary sheet stays accurate even when
        # self.asins is later narrowed down to a retry-only subset.
        self.total_asins_in_csv = len(self.asins)

    def _load_asins_from_csv(self):
        """Load ASINs from CSV file"""
        try:
            with open(self.csv_file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    # Skip headers and empty rows
                    if row and row[0].strip() and not row[0].startswith('Brand') and not row[0].startswith('Row') and row[0].strip() != 'Grand Total':
                        asin = row[0].strip()
                        # Check if it looks like a valid ASIN (starts with B and is alphanumeric)
                        if asin.startswith('B') and len(asin) >= 10:
                            self.asins.append(asin)

            print(f"✓ Loaded {len(self.asins)} ASINs from CSV file")
            return self.asins
        except FileNotFoundError:
            print(f"❌ CSV file not found: {self.csv_file_path}")
            return []
        except Exception as e:
            print(f"⚠️ Error loading CSV: {str(e)}")
            return []

    def _init_selenium(self):
        """Initialize Selenium WebDriver for scraping seller info"""
        try:
            options = webdriver.ChromeOptions()
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')

            self.driver = webdriver.Chrome(options=options)
            self.driver.set_page_load_timeout(30)
            print("✓ Selenium WebDriver initialized")
        except Exception as e:
            print(f"⚠️ Warning: Could not initialize Selenium: {str(e)}")
            print("  Seller info will be set to 'Not Available'")
            self.driver = None

    def _force_kill_driver(self):
        """
        Kill the chromedriver/Chrome process tree directly instead of calling
        driver.quit(), which sends an HTTP command through the same channel
        that may be the thing that's wedged.
        """
        if self.driver:
            try:
                pid = self.driver.service.process.pid
            except Exception:
                pid = None
            if pid:
                try:
                    if platform.system() == "Windows":
                        subprocess.run(
                            ["taskkill", "/F", "/T", "/PID", str(pid)],
                            capture_output=True, timeout=10
                        )
                    else:
                        import signal
                        os.kill(pid, signal.SIGKILL)
                except Exception:
                    pass
        self.driver = None

    def _restart_selenium(self):
        """Recover from a wedged browser: force-kill it and start a fresh session."""
        self._force_kill_driver()
        self._init_selenium()

    def extract_seller_info_from_page(self, asin):
        """
        Extract Fulfilled by, Sold by, and Shipper/Seller information from Amazon.ae product page
        Returns: (fulfilled_by, sold_by, shipper_seller)
        """
        if not self.driver:
            return "Not Available", "Not Available", "Not Available"

        def _extract():
            url = f"https://www.amazon.ae/dp/{asin}"
            self.driver.get(url)

            # Wait for page to load
            WebDriverWait(self.driver, config.WAIT_TIMEOUT).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(1)

            # Scroll to load dynamic content
            self.driver.execute_script("window.scrollBy(0, 2000);")
            time.sleep(0.5)

            # Get page text
            page_text = self.driver.find_element(By.TAG_NAME, "body").text
            lines = page_text.split('\n')

            fulfilled_by = "Not Available"
            sold_by = "Not Available"
            shipper_seller = "Not Available"

            # Parse seller information from page text
            for i, line in enumerate(lines):
                line_stripped = line.strip()
                next_line = lines[i + 1].strip() if i + 1 < len(lines) else ''

                # Look for "Fulfilled by"
                if 'Fulfilled by' in line_stripped and fulfilled_by == "Not Available":
                    if ':' in line_stripped:
                        fulfilled_by = line_stripped.split(':', 1)[1].strip()[:100]
                    elif len(line_stripped) > len('Fulfilled by'):
                        fulfilled_by = line_stripped.replace('Fulfilled by', '').strip()[:100]
                    else:
                        fulfilled_by = next_line[:100] if next_line else "Not Available"

                # Look for "Sold by"
                if 'Sold by' in line_stripped and sold_by == "Not Available":
                    if ':' in line_stripped:
                        sold_by = line_stripped.split(':', 1)[1].strip()[:100]
                    elif len(line_stripped) > len('Sold by'):
                        sold_by = line_stripped.replace('Sold by', '').strip()[:100]
                    else:
                        sold_by = next_line[:100] if next_line else "Not Available"

                # Look for "Shipper / Seller" (with space variations)
                if 'Shipper' in line_stripped and '/' in line_stripped and 'Seller' in line_stripped:
                    if shipper_seller == "Not Available":
                        if ':' in line_stripped:
                            shipper_seller = line_stripped.split(':', 1)[1].strip()[:100]
                        elif len(line_stripped) > len('Shipper / Seller'):
                            shipper_seller = line_stripped.replace('Shipper / Seller', '').replace('Shipper/Seller', '').strip()[:100]
                        else:
                            shipper_seller = next_line[:100] if next_line else "Not Available"

            outcome['value'] = (fulfilled_by, sold_by, shipper_seller)

        outcome = {}

        def _run():
            try:
                _extract()
            except Exception as e:
                outcome['error'] = str(e)

        worker = threading.Thread(target=_run, daemon=True)
        worker.start()
        worker.join(timeout=SELENIUM_TIMEOUT)

        if worker.is_alive():
            print(f"  ⚠️ Selenium wedged on {asin} (no response after {SELENIUM_TIMEOUT}s), restarting browser...")
            self._restart_selenium()
            return "Not Available", "Not Available", "Not Available"

        if 'error' in outcome:
            print(f"  ⚠️ Error extracting seller info for {asin}: {outcome['error']}")
            # A crashed/disconnected browser leaves self.driver pointing at a
            # dead session -- every subsequent ASIN would silently fail the
            # same way for the rest of the run unless we restart here too
            # (not just on the hang/timeout path above).
            error_lower = outcome['error'].lower()
            dead_session_markers = (
                "invalid session id", "chrome not reachable", "disconnected",
                "session deleted", "no such window", "target window already closed",
                "unable to connect", "connection refused",
            )
            if any(marker in error_lower for marker in dead_session_markers):
                print(f"  ⚠️ Browser session appears dead, restarting...")
                self._restart_selenium()
            return "Not Available", "Not Available", "Not Available"

        return outcome.get('value', ("Not Available", "Not Available", "Not Available"))

    def fetch_product_details(self, asin):
        """Fetch product details from Rainforest API, retrying transient network errors"""
        params = {
            "api_key": self.api_key,
            "amazon_domain": "amazon.ae",
            "type": "product",
            "asin": asin
        }

        response = None
        for attempt in range(1, config.MAX_RETRIES + 1):
            try:
                response = requests.get(self.base_url, params=params, timeout=30)
                response.raise_for_status()
                break
            except requests.exceptions.RequestException as e:
                if attempt < config.MAX_RETRIES:
                    print(f"  ⚠️ Network error for {asin} (attempt {attempt}/{config.MAX_RETRIES}): {str(e)}, retrying...")
                    time.sleep(2)
                else:
                    print(f"  ⚠️ Network error for {asin}: {str(e)}")
                    return None

        try:
            data = response.json()

            if not data.get("request_info", {}).get("success"):
                print(f"  ⚠️ API Error for {asin}: {data.get('request_info', {}).get('error_message', 'Unknown error')}")
                return None

            product = data.get("product", {})
            if not product:
                print(f"  ⚠️ No product data for {asin}")
                return None

            # Extract price info (Rainforest API returns the live price under
            # buybox_winner.price, not a top-level "prices" list). Use the
            # numeric "value" rather than "raw" so Price stays a sortable
            # number in Excel instead of duplicating the Currency column
            # (raw looks like "AED1,031.00"). Some ASINs genuinely have no
            # live buybox price (out of stock / no qualifying offer) -- N/A
            # in that case reflects Rainforest's data, not a parsing bug.
            price_info = product.get("buybox_winner", {}).get("price") or product.get("price") or {}
            price = price_info.get("value", "N/A")

            # Extract brand (top-level field, falling back to the specifications table)
            brand = product.get("brand")
            if not brand:
                for spec in product.get("specifications", []):
                    if spec.get("name") in ("Brand", "Brand Name"):
                        brand = spec.get("value")
                        break
            brand = brand or "N/A"

            # Extract seller info from page using Selenium
            print(f"    Extracting seller info for {asin}...", end=" ")
            fulfilled_by, sold_by, shipper_seller = self.extract_seller_info_from_page(asin)
            print(f"✓")

            product_info = {
                "Brand": brand,
                "ASIN": asin,
                "Title": product.get("title", "N/A"),
                "Price": price,
                "Currency": price_info.get("currency", "AED"),
                "Fulfilled by": fulfilled_by,
                "Sold by": sold_by,
                "Shipper/Seller": shipper_seller,
                "Specs": product.get("specifications_flat") or product.get("description", ""),
                "Rating": product.get("rating", "N/A"),
                "Reviews": product.get("ratings_total", 0),
                "URL": f"https://www.amazon.ae/dp/{asin}"
            }

            return product_info

        except requests.exceptions.RequestException as e:
            print(f"  ⚠️ Network error for {asin}: {str(e)}")
            return None
        except Exception as e:
            print(f"  ⚠️ Error fetching {asin}: {str(e)}")
            return None

    def collect_all_products(self):
        """Collect products for all ASINs from CSV"""
        print("\n" + "="*70)
        print("🚀 AMAZON.AE VACUUM CLEANER DATA COLLECTION")
        print("   (Using filtered ASINs from CSV + Enhanced seller info extraction)")
        print("="*70)
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"API Key: {self.api_key[:10]}...")
        print(f"Total ASINs to process: {len(self.asins)}\n")

        for idx, asin in enumerate(self.asins, 1):
            print(f"[{idx}/{len(self.asins)}] Processing {asin}...", end=" ")
            product = self.fetch_product_details(asin)

            if product:
                self.all_products.append(product)
                print(f"✓ {product['Title'][:40]}...")
            else:
                print("✗ Failed")

            # Respectful delay to avoid rate limiting
            time.sleep(config.REQUEST_DELAY)

        print(f"\n✅ Total products collected: {len(self.all_products)}/{len(self.asins)}")
        return self.all_products

    @staticmethod
    def load_products_from_excel(xlsx_path):
        """
        Read an existing output file's "All Products" sheet back into a list
        of product dicts, keyed by the same headers create_excel_file writes.
        Used by --retry-failed to merge newly-recovered ASINs into a prior run
        without re-fetching everything that already succeeded.
        """
        products = []
        if not os.path.exists(xlsx_path):
            return products

        wb = load_workbook(xlsx_path)
        if "All Products" not in wb.sheetnames:
            return products

        ws = wb["All Products"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            products.append(dict(zip(headers, row)))
        return products

    def create_excel_file(self, output_filename="Amazon_Vacuum_Cleaners_Filtered.xlsx"):
        """Create Excel workbook with product data"""
        print(f"\n📝 Creating Excel file: {output_filename}")

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create Summary Sheet
        summary_sheet = wb.create_sheet("Summary", 0)
        summary_sheet['A1'] = "Data Collection Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        summary_sheet['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        summary_sheet.merge_cells('A1:D1')

        summary_sheet['A3'] = "Metric"
        summary_sheet['B3'] = "Value"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ['A3', 'B3']:
            summary_sheet[cell].fill = header_fill
            summary_sheet[cell].font = header_font

        summary_data = [
            ["Total Products Collected", len(self.all_products)],
            ["Total ASINs in CSV", self.total_asins_in_csv],
            ["Collection Date", datetime.now().strftime('%Y-%m-%d')],
            ["Collection Time", datetime.now().strftime('%H:%M:%S')],
            ["Avg Rating", f"{sum([float(p['Rating']) if isinstance(p['Rating'], (int, float)) else 0 for p in self.all_products]) / len(self.all_products) if self.all_products else 0:.1f}"]
        ]

        for i, (metric, value) in enumerate(summary_data, start=4):
            summary_sheet[f'A{i}'] = metric
            summary_sheet[f'B{i}'] = value

        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 25

        # Create All Products Sheet
        products_sheet = wb.create_sheet("All Products", 1)

        headers = ["Brand", "ASIN", "Title", "Price", "Currency", "Fulfilled by", "Sold by", "Shipper/Seller", "Specs", "Rating", "Reviews", "URL"]
        products_sheet.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in products_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add product data
        for product in self.all_products:
            products_sheet.append([
                product["Brand"],
                product["ASIN"],
                product["Title"],
                product["Price"],
                product["Currency"],
                product["Fulfilled by"],
                product["Sold by"],
                product["Shipper/Seller"],
                product["Specs"],
                product["Rating"],
                product["Reviews"],
                product["URL"]
            ])

        # Set column widths
        products_sheet.column_dimensions['A'].width = 15
        products_sheet.column_dimensions['B'].width = 12
        products_sheet.column_dimensions['C'].width = 40
        products_sheet.column_dimensions['D'].width = 12
        products_sheet.column_dimensions['E'].width = 10
        products_sheet.column_dimensions['F'].width = 20
        products_sheet.column_dimensions['G'].width = 20
        products_sheet.column_dimensions['H'].width = 20
        products_sheet.column_dimensions['I'].width = 50
        products_sheet.column_dimensions['J'].width = 10
        products_sheet.column_dimensions['K'].width = 10
        products_sheet.column_dimensions['L'].width = 45

        # Freeze header row
        products_sheet.freeze_panes = "A2"

        # Save workbook
        wb.save(output_filename)
        print(f"✅ Excel file created successfully: {output_filename}")
        return output_filename

    def close(self):
        """Clean up Selenium driver, force-killing it if quit() itself hangs"""
        if not self.driver:
            return
        worker = threading.Thread(target=self.driver.quit, daemon=True)
        worker.start()
        worker.join(timeout=15)
        if worker.is_alive():
            self._force_kill_driver()
        print("\n✓ Browser closed")

def main():
    # Configuration (from .env / environment via config.py, with fallbacks)
    API_KEY = config.RAINFOREST_API_KEY
    CSV_FILE_PATH = config.CSV_FILE_PATH
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(config.OUTPUT_DIR, config.OUTPUT_FILENAME)

    # --retry-failed: instead of re-fetching all 108 ASINs, only fetch the
    # ones missing from the existing output file (whether they hard-failed
    # or the run was interrupted before reaching them), then merge into it.
    retry_mode = "--retry-failed" in sys.argv

    collector = None
    try:
        collector = AmazonVacuumCollector(API_KEY, CSV_FILE_PATH)

        if not collector.asins:
            print("\n❌ No ASINs loaded from CSV file. Exiting.")
            return

        if retry_mode:
            existing_products = collector.load_products_from_excel(output_path)
            existing_asins = {p.get("ASIN") for p in existing_products}
            missing_asins = [a for a in collector.asins if a not in existing_asins]

            if not missing_asins:
                print(f"\n✅ All {collector.total_asins_in_csv} ASINs already present in {output_path} -- nothing to retry.")
                return

            print(f"\n🔁 Retry mode: {len(existing_products)} product(s) already in {output_path}, retrying {len(missing_asins)} missing ASIN(s)...")
            collector.asins = missing_asins
            collector.collect_all_products()

            collector.all_products = existing_products + collector.all_products
            output_file = collector.create_excel_file(output_path)
            recovered = len(collector.all_products) - len(existing_products)
            print(f"\n🎉 Retry complete. Recovered {recovered}/{len(missing_asins)}.")
            print(f"   Total products now: {len(collector.all_products)}/{collector.total_asins_in_csv}. File: {output_file}")
            return

        collector.collect_all_products()

        if collector.all_products:
            output_file = collector.create_excel_file(output_path)
            print(f"\n🎉 SUCCESS! File ready: {output_file}")
            print("="*70)
            print("Features:")
            print("  ✅ Data from filtered ASINs in your CSV")
            print("  ✅ Enhanced seller info (Fulfilled by, Sold by, Shipper/Seller)")
            print("  ✅ 3-column seller info extraction")
            print("  ✅ Summary sheet with collection stats")
            print("="*70)
            missing = collector.total_asins_in_csv - len(collector.all_products)
            if missing:
                print(f"  ⚠️ {missing} ASIN(s) failed to collect -- rerun with --retry-failed to fill the gaps")
        else:
            print("\n❌ No products were collected. Please check your CSV file and API key.")

    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        print("Please ensure you have:")
        print("  - Internet access")
        print("  - Valid Rainforest API key")
        print("  - CSV file with ASINs in the same directory")
        print("  - Google Chrome installed")
        print("  - ChromeDriver installed or in PATH")

    finally:
        if collector:
            collector.close()

if __name__ == "__main__":
    main()